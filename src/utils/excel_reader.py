from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from collections import namedtuple
from utils import config_reader
from utils import constants


Col = namedtuple("Col", "location name")

def get_mapping_sheet_data(config, excel_path: str):

    workbook: Workbook = load_workbook(
        filename=excel_path, read_only=True, data_only=True)

    target_table_name = config['Target']['TableName']
    sheet_data_dict = {}

    sheet_data_dict['rules_data'] = get_raw_sheet_data(
        workbook[config['Rules']['Sheet']], config_reader.rules_headers, 'Rules', config)
    sheet_data_dict['mapping_data'] = get_raw_sheet_data(
        workbook[config['Mapping']['Sheet']], config_reader.mapping_headers, 'Mapping', config)

    merged_cells_fix(sheet_data_dict['mapping_data'][constants.TRG_COL_NAME_VAL])
    merged_cells_fix(sheet_data_dict['mapping_data'][constants.TRG_COL_DESCR_VAL])
    merged_cells_fix(sheet_data_dict['mapping_data'][constants.TRG_COL_DATATYPE_VAL])
    merged_cells_fix(sheet_data_dict['mapping_data'][constants.TRG_COL_MODE_VAL])
    merged_cells_fix(sheet_data_dict['mapping_data'][constants.TRG_COL_SENSITIVE_FLAG])

    return (sheet_data_dict, target_table_name)

"""
looks at a given sheet and specified columns for scanning
returns dictionary with columns and their rows
"""


def get_raw_sheet_data(sheet, column_header_loc_list, map_name_sheet, config):
    """
    Returns coordinate and value for column header in the sheet
    """

    def get_header_information(sheet, coordinate):
        return Col(coordinate, sheet[coordinate].value)

    def get_rows_of_column(sheet, header_alphanum_coordinate):
        rows = []
        xy = coordinate_from_string(header_alphanum_coordinate)
        col_idx = column_index_from_string(xy[0])
        #max_row_iter = int(config['MappingEndScan'][map_name_sheet])

        for row in sheet.iter_rows(min_row=xy[1] + 1, min_col=col_idx, max_col=col_idx):
            rows.append(row[0].value)

        return rows

    dataDict = {}

    for column_header_reference in column_header_loc_list:
        
        header = get_header_information(
            sheet, config[map_name_sheet][column_header_reference])

        dataDict[column_header_reference] = get_rows_of_column(sheet, header.location)

    return dataDict

"""
Arguments: lists
Merged cells keep only cell with most value. other sub-cells are None 

specify which column needs fix for merged cells
this will scan column upwards for every cell which has no value
scanning will happen until first non-empty value is found

To make scan work correctly a cell must be given a reference column 
the scan will stop once index of scanning applied to reference column will give value different from original reference cell
That will ensure that cells which have intentional empty values will stay empty

ref_col when None, assumes there are no intentional "None" cells, so it will always fill it with value
"""


def merged_cells_fix(fix_col, ref_col=None):
    idx = 0

    for row in fix_col:

        if ref_col is not None:
            orig_ref_row_val = ref_col[idx]

        if row is None and idx > 0:

            go_back_idx = 1
            prev_row_val = None

            while idx - go_back_idx >= 0:
                prev_row_val = fix_col[idx - go_back_idx]
                if (prev_row_val is not None):
                    if (ref_col is not None and ref_col[idx - go_back_idx] != orig_ref_row_val):
                        break
                    else:
                        fix_col[idx] = prev_row_val
                        break
                else:
                    go_back_idx += 1
        idx += 1
